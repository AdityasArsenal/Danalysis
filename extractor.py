from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt

wb  = load_workbook('ee.xlsx')
ws = wb.active

print(ws['B469'].value)

ll = [467,468]
yr = []
amt = []

for i in ll:
    amt.append(ws['F'+str(i)].value)
    yr.append(ws['C'+str(i)].value)

print(amt)
print(yr)




# Example data
x = yr  # X-axis labels (months)
y = amt # Y-axis values (data points)

# Plot the graph
plt.plot(x, y, marker='o', linestyle='-', color='b')

# Add labels and title
plt.xlabel("Periods")
plt.ylabel("Values")
plt.title("Yearly Data")

# Show the graph
plt.show()
