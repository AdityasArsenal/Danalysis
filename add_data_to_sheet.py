from openpyxl import load_workbook, Workbook
import os

def adder(KPIs, Values, Referance_Unit, Units, Periods, Decimals, Not_Found, name_of_sheet):
    """
    Adds data to a new sheet in the Excel workbook.
    If the file doesn't exist, creates a new one.
    """
    file_path = "res.xlsx"
    
    # Create new workbook if it doesn't exist
    if not os.path.exists(file_path):
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    else:
        wb = load_workbook(file_path)
    
    # Generate unique sheet name
    base_name = name_of_sheet
    sheet_name = base_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_name}_{counter}"
        counter += 1
    
    # Create new sheet
    new_sheet = wb.create_sheet(sheet_name)
    
    # Add headers
    headers = ['Name', 'KPI', 'Value', 'Unit', 'Period', 'Decimal', 'Referance_Unit', 'Not_Found']
    for col, header in enumerate(headers, 1):
        new_sheet.cell(row=1, column=col, value=header)
    
    # Add data for found KPIs
    for i, (kpi, value, ref_unit, unit, period, decimal) in enumerate(
        zip(KPIs, Values, Referance_Unit, Units, Periods, Decimals), 2):
        new_sheet.cell(row=i, column=2, value=kpi)
        new_sheet.cell(row=i, column=3, value=value)
        new_sheet.cell(row=i, column=4, value=ref_unit)
        new_sheet.cell(row=i, column=5, value=unit)
        new_sheet.cell(row=i, column=6, value=period)
        new_sheet.cell(row=i, column=7, value=decimal)
    
    # Add Not Found KPIs in the last column
    for i, not_found_kpi in enumerate(Not_Found, 2):
        if not_found_kpi != "None":  # Skip "None" values
            new_sheet.cell(row=i, column=8, value=not_found_kpi)
    
    # Save the workbook
    wb.save(file_path)
    print(f"Data added to new sheet: {sheet_name}")
    
        