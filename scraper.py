import requests
import os
from openpyxl import load_workbook

def download_xml_files(file_path) -> None:
    os.makedirs('new_xml', exist_ok=True)
    headers = {



    

    
    
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    wb = load_workbook(file_path)
    ws = wb.active
    
    company_names = []
    urls = []
    file_names = []

    j=1
    for i in range(1,len(ws["A"])): 

        company_names.append((ws["A"][i].value))
        urls.append((ws["E"][i].value))
        j+=1


    for i, url in enumerate(urls, 0):
        if i > 10:                                             
          break
        try:
            filename = os.path.join('new_xml', f"{company_names[i]}_.xml")
            file_names.append(filename)

            print(f"Downloading file {i} of 10...")
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            
            # Save in new directory 
            with open(filename, "wb") as f:
                f.write(response.content)
            
        except requests.RequestException as e:
            print(f"❌ Error downloading {url}: {str(e)}")
    
    return company_names,file_names,urls
